#!/usr/bin/env python3
"""
无 LibreOffice 时的兜底"重算"——纯 Python 计算费用明细表里的简单公式并核对。

为什么需要它:
  SKILL 推荐用 xlsx skill 的 recalc.py(基于 LibreOffice)把公式算出缓存值。
  但在没有 LibreOffice、或没有 /mnt/skills 路径的本地/精简环境里 recalc.py 跑不了。
  本模板里的公式只有三种(=SUM(区间)、=单元格-单元格、=单元格引用),
  本脚本对这三种做求值, 既能核对(--check, 默认)也能把数值写回(--write)。

用法:
  python3 recalc_fallback.py <模板副本.xlsx>            # 计算并打印各列小计/总额(不改文件)
  python3 recalc_fallback.py <模板副本.xlsx> --write    # 把公式结果写回单元格(数值可见, 适合交付)

输出 JSON: total_formulas / total_errors / values(关键单元格算得的数值)。
total_errors>0 表示有看不懂的公式, 需人工检查。
"""
import argparse, json, re, sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

NUM = (int, float)

def to_num(v):
    if isinstance(v, NUM):
        return float(v)
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", "").replace("¥", "").replace("￥", "").strip())
    except ValueError:
        return None  # 文本, 当 0 处理但记一笔

def eval_formula(f, getval):
    """支持 =SUM(a:b)[+SUM(c:d)...]、=A1-B1、=A1。返回(数值, 是否成功)。"""
    s = f[1:].replace(" ", "")
    # 纯 SUM 之和: SUM(...)+SUM(...)
    if s.upper().startswith("SUM("):
        total = 0.0
        for m in re.finditer(r"SUM\(([^)]+)\)", s, re.I):
            rng = m.group(1)
            try:
                c1, r1, c2, r2 = range_boundaries(rng)
            except Exception:
                return None, False
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    n = to_num(getval(r, c))
                    total += (n or 0.0)
        return round(total, 2), True
    # 形如 H40-M40
    m = re.fullmatch(r"([A-Z]+\d+)-([A-Z]+\d+)", s, re.I)
    if m:
        a, b = ref_num(m.group(1), getval), ref_num(m.group(2), getval)
        return round(a - b, 2), True
    # 形如 =R40 单引用
    m = re.fullmatch(r"([A-Z]+\d+)", s, re.I)
    if m:
        return round(ref_num(m.group(1), getval), 2), True
    return None, False

def ref_num(ref, getval):
    m = re.fullmatch(r"([A-Z]+)(\d+)", ref, re.I)
    from openpyxl.utils import column_index_from_string
    c = column_index_from_string(m.group(1).upper()); r = int(m.group(2))
    return to_num(getval(r, c)) or 0.0

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--write", action="store_true", help="把公式结果写回单元格")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx)
    ws = wb.active
    # 当前快照(公式单元格在解出前按 0 处理, 多轮迭代直到稳定)
    computed = {}  # (r,c) -> number

    def getval(r, c):
        if (r, c) in computed:
            return computed[(r, c)]
        return ws.cell(row=r, column=c).value

    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((cell.row, cell.column, cell.value))

    errors = []
    # 迭代求值(总额依赖小计, 故多跑几轮)
    for _ in range(5):
        changed = False
        for r, c, f in formula_cells:
            val, ok = eval_formula(f, getval)
            if ok and computed.get((r, c)) != val:
                computed[(r, c)] = val; changed = True
        if not changed:
            break
    for r, c, f in formula_cells:
        if (r, c) not in computed:
            errors.append({"cell": f"{get_column_letter(c)}{r}", "formula": f})

    if args.write:
        for (r, c), v in computed.items():
            ws.cell(row=r, column=c).value = v
        wb.save(args.xlsx)

    # 摘要: 报销总额行/费用报销金额(尽力定位)
    def label_row(lbl):
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(row=r, column=1).value).strip() == lbl:
                return r
        return None
    tr = label_row("费用报销总额")
    summary = {}
    if tr:
        from openpyxl.utils import column_index_from_string as ci
        summary["费用合计H"] = computed.get((tr, ci("H")))
        summary["费用报销金额R"] = computed.get((tr, ci("R")))

    print(json.dumps({
        "status": "ok" if not errors else "has_unparsed_formula",
        "xlsx": args.xlsx, "written": bool(args.write),
        "total_formulas": len(formula_cells), "total_errors": len(errors),
        "unparsed": errors, "summary": summary,
    }, ensure_ascii=False, indent=2))
    sys.exit(0 if not errors else 1)

if __name__ == "__main__":
    main()
