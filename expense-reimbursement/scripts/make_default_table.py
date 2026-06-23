#!/usr/bin/env python3
"""
没有用户模板时, 用本脚本从 ledger.json 生成一份通用「费用明细表」(xlsx + md)。

用法:
  python3 make_default_table.py --ledger 过程文件/ledger.json --out <工作区>/费用明细表.xlsx

ledger.json (无模板版) 结构:
{
  "identity": {                # 都可选, 有就写进表头
    "标题": "5月AI报销费用明细",
    "报销人": "叶守淦", "部门": "产品部",
    "公司主体": "深圳银河通用机器人有限公司",
    "提交日期": "2026-06-22", "币种": "RMB", "预支现金": 0
  },
  "items": [
    { "日期":"2026-05-07", "类别":"软件订阅", "事由":"Suno Pro 月度订阅",
      "商户":"Suno", "币种":"USD", "原金额":10, "汇率":7.10, "金额":71.00, "备注":"" }
    # "金额" 为最终入表 RMB 数。币种=RMB 时 原金额/汇率 可留空。
  ]
}

跨平台: 纯 python3 + openpyxl, 不依赖 LibreOffice / mac / linux 专有功能。
输出 xlsx 直接写好数值与合计(不留公式), 任何环境打开都显示数字; 同时输出同名 .md 便于预览。
"""
import argparse, json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

COLS = ["序号", "日期", "费用类别", "事由/说明", "商户", "币种", "原金额", "汇率", "金额(元)", "备注"]

def num(v):
    return v if isinstance(v, (int, float)) else None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ledger", required=True)
    ap.add_argument("--out", required=True, help="输出 xlsx 路径(同名 .md 一并生成)")
    args = ap.parse_args()

    with open(args.ledger, encoding="utf-8") as f:
        data = json.load(f)
    ident = data.get("identity", {})
    items = data.get("items", [])

    # 按日期升序(缺日期排最后), 稳定
    items = sorted(items, key=lambda it: (str(it.get("日期") or "9999")))

    wb = Workbook(); ws = wb.active; ws.title = "费用明细表"
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="F2F2F2")

    title = ident.get("标题") or "费用明细表"
    ws.append([title]); ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    ws["A1"].alignment = Alignment(horizontal="center")

    # 身份信息行(只写存在的字段)
    id_pairs = [(k, ident[k]) for k in ("报销人", "部门", "公司主体", "提交日期", "币种") if ident.get(k) not in (None, "")]
    if id_pairs:
        ws.append([" ｜ ".join(f"{k}: {v}" for k, v in id_pairs)])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(COLS))

    # 表头
    head_row = ws.max_row + 1
    ws.append(COLS)
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(row=head_row, column=c)
        cell.font = Font(bold=True); cell.fill = hdr_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center")

    total = 0.0
    for i, it in enumerate(items, 1):
        amt = num(it.get("金额"))
        if amt is not None:
            total += amt
        ws.append([
            i, it.get("日期", ""), it.get("类别", ""), it.get("事由", ""),
            it.get("商户", ""), it.get("币种", "RMB"),
            num(it.get("原金额")), num(it.get("汇率")),
            amt, it.get("备注", ""),
        ])
        for c in range(1, len(COLS) + 1):
            ws.cell(row=ws.max_row, column=c).border = border

    total = round(total, 2)
    adv = num(ident.get("预支现金")) or 0
    ws.append(["", "", "", "合计", "", "", "", "", total, ""])
    tr = ws.max_row
    for c in range(1, len(COLS) + 1):
        ws.cell(row=tr, column=c).border = border
        ws.cell(row=tr, column=c).font = Font(bold=True)
    if adv:
        ws.append(["", "", "", f"减预支现金 {adv}", "", "", "", "", round(total - adv, 2), ""])

    widths = [5, 12, 12, 34, 16, 6, 9, 7, 11, 24]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    os.makedirs(os.path.dirname(os.path.abspath(args.out)) or ".", exist_ok=True)
    wb.save(args.out)

    # 同名 markdown
    md_path = os.path.splitext(args.out)[0] + ".md"
    lines = [f"# {title}", ""]
    if id_pairs:
        lines.append(" ｜ ".join(f"**{k}**: {v}" for k, v in id_pairs)); lines.append("")
    lines.append("| " + " | ".join(COLS) + " |")
    lines.append("|" + "|".join(["---"] * len(COLS)) + "|")
    for i, it in enumerate(items, 1):
        row = [str(i), it.get("日期", ""), it.get("类别", ""), it.get("事由", ""),
               it.get("商户", ""), it.get("币种", "RMB"),
               str(it.get("原金额", "") or ""), str(it.get("汇率", "") or ""),
               str(num(it.get("金额")) or ""), str(it.get("备注", "") or "")]
        lines.append("| " + " | ".join(s.replace("|", "/") for s in row) + " |")
    lines.append("| | | | **合计** | | | | | **%.2f** | |" % total)
    # 用 UTF-8 带 BOM 写 md, 保证 Windows 记事本/Excel 打开标题与中文不乱码。
    with open(md_path, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(lines) + "\n")

    print(json.dumps({
        "status": "ok", "xlsx": args.out, "md": md_path,
        "items": len(items), "total": total,
        "reimburse": round(total - adv, 2),
    }, ensure_ascii=False))

if __name__ == "__main__":
    main()
