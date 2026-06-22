#!/usr/bin/env python3
"""
把整理好的费用台账回填进【模板副本】(原模板绝不修改)。

用法:
  python fill_template.py --xlsx <模板副本.xlsx> --ledger <费用台账.json>

费用台账 JSON 结构:
{
  "identity": {
    "公司主体": "北京银河通用机器人有限公司",   # 写入 C3, 缺省保留模板原值
    "部门": "开发者生态部",                      # G3
    "姓名": "叶守淦",                            # J3
    "提交日期": "2026-06-22",                    # M3 (字符串原样写入)
    "币种": "RMB",                               # R3
    "预支现金": 0,                               # M40 (总额行)
    "默认所属部门": "开发者生态部",              # 行级 F 列缺省值
    "默认所属项目编码": "无"                     # 行级 G 列缺省值
  },
  "items": [
    {
      "月": 6, "日": 15,
      "事由": "北京-上海出差 火车票(高铁二等座)",
      "起": "北京", "止": "上海",          # 写入 D / E (非差旅类可留空)
      "所属部门": "开发者生态部",          # 缺省取 identity.默认所属部门
      "所属项目编码": "无",                # 缺省取 identity.默认所属项目编码
      "列": "H",                           # 直接指定模板列(H..S), 优先级最高
      "类别": "火车票机票",                # 或给类别名, 由下方映射换算成列
      "金额": 553.5                        # 最终 RMB 金额(外币请先换算)
    }
  ]
}

字段优先级: 若给了 "列" 用之; 否则用 "类别" 查映射表。两者都缺则报错。
回填后金额公式不会自动算出数值, 请随后运行 xlsx skill 的 recalc.py 重算。
"""
import argparse, json, sys
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# 类别名 -> 模板列。键尽量覆盖常见说法。
CATEGORY_TO_COL = {
    "火车票": "H", "机票": "H", "火车票机票": "H", "火车票/机票": "H", "高铁": "H", "飞机票": "H",
    "住宿": "I", "住宿费": "I", "酒店": "I",
    "交通费": "J", "差旅交通": "J", "差旅交通费": "J", "出差交通": "J",
    "出差补助": "K", "补助": "K", "差旅补助": "K",
    "办公费": "L", "办公": "L", "办公用品": "L",
    "福利费": "M", "福利": "M", "团建": "M", "零食": "M",
    "工作餐": "N", "餐费": "N", "误餐": "N",
    "研发物料": "O", "物料": "O",
    "招待费": "P", "招待": "P", "业务招待": "P",
    "市内交通费": "Q", "市内交通": "Q", "打车": "Q", "出租车": "Q", "地铁": "Q",
    "会议费": "R", "培训费": "R", "会议费培训费": "R", "会议/培训": "R",
    "其他": "S",
}
DATA_START = 8

def find_label_row(ws, label):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value).strip() == label:
            return r
    return None

def resolve_col(item):
    if item.get("列"):
        c = str(item["列"]).strip().upper()
        if c not in [get_column_letter(i) for i in range(column_index_from_string("H"), column_index_from_string("S") + 1)]:
            raise ValueError(f"列 {c} 超出金额列范围 H..S")
        return c
    cat = str(item.get("类别", "")).strip()
    if cat in CATEGORY_TO_COL:
        return CATEGORY_TO_COL[cat]
    raise ValueError(f"无法确定列: 条目缺少有效的 '列' 或可识别的 '类别' (得到类别={cat!r}, 事由={item.get('事由')!r})")

def copy_row_style(ws, src_row, dst_row, cols):
    for c in range(1, cols + 1):
        s = ws.cell(row=src_row, column=c)
        d = ws.cell(row=dst_row, column=c)
        if s.has_style:
            d.font = copy(s.font); d.border = copy(s.border)
            d.fill = copy(s.fill); d.alignment = copy(s.alignment)
            d.number_format = s.number_format

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--ledger", required=True)
    ap.add_argument("--no-sort", action="store_true",
                    help="不按时间排序(默认按 月/日 升序排, 因模板要求按时间顺序填写)")
    args = ap.parse_args()

    with open(args.ledger, encoding="utf-8") as f:
        data = json.load(f)
    ident = data.get("identity", {})
    items = data.get("items", [])

    # 模板表头批注要求"按时间顺序依次填写"。默认按 (月,日) 稳定升序排序;
    # 同日内保留 ledger 原顺序(可用此控制同日多笔的先后, 如先正票后退票)。
    if not args.no_sort:
        items = sorted(items, key=lambda it: (
            it.get("月") if isinstance(it.get("月"), (int, float)) else 99,
            it.get("日") if isinstance(it.get("日"), (int, float)) else 99,
        ))

    wb = load_workbook(args.xlsx)
    ws = wb.active

    subtotal_row = find_label_row(ws, "小计")
    total_row = find_label_row(ws, "费用报销总额")
    if not subtotal_row or not total_row:
        print("ERROR: 模板中找不到 '小计' 或 '费用报销总额' 行, 模板结构异常", file=sys.stderr); sys.exit(2)

    available = subtotal_row - DATA_START
    n = len(items)
    if n > available:
        k = n - available
        ws.insert_rows(subtotal_row, amount=k)
        for r in range(subtotal_row, subtotal_row + k):
            copy_row_style(ws, DATA_START, r, 19)
        subtotal_row = find_label_row(ws, "小计")
        total_row = find_label_row(ws, "费用报销总额")

    last_data_row = subtotal_row - 1
    # 拆掉落在数据区内的残留合并(否则 MergedCell 只读写不进)
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= DATA_START and mc.max_row <= last_data_row:
            ws.unmerge_cells(str(mc))
    # 清空数据区(含模板自带的示例数字)
    for r in range(DATA_START, subtotal_row):
        for c in range(1, 20):
            ws.cell(row=r, column=c).value = None

    # 写身份信息(合并单元格写左上锚点即可)
    def setv(coord, val):
        if val is not None and val != "":
            ws[coord] = val
    setv("C3", ident.get("公司主体"))
    setv("G3", ident.get("部门"))
    setv("J3", ident.get("姓名"))
    setv("M3", ident.get("提交日期"))
    setv("R3", ident.get("币种"))

    def_dept = ident.get("默认所属部门")
    def_proj = ident.get("默认所属项目编码")

    # 写明细行
    for i, it in enumerate(items):
        r = DATA_START + i
        col = resolve_col(it)
        ws.cell(row=r, column=1).value = it.get("月")
        ws.cell(row=r, column=2).value = it.get("日")
        ws.cell(row=r, column=3).value = it.get("事由")
        ws.cell(row=r, column=4).value = it.get("起")
        ws.cell(row=r, column=5).value = it.get("止")
        ws.cell(row=r, column=6).value = it.get("所属部门") or def_dept
        ws.cell(row=r, column=7).value = it.get("所属项目编码") or def_proj
        ws.cell(row=r, column=column_index_from_string(col)).value = it.get("金额")

    # 重写小计 SUM (H..S)
    for c in range(column_index_from_string("H"), column_index_from_string("S") + 1):
        L = get_column_letter(c)
        ws.cell(row=subtotal_row, column=c).value = f"=SUM({L}{DATA_START}:{L}{last_data_row})"

    # 重写总额行公式 + 预支现金
    ws.cell(row=total_row, column=column_index_from_string("H")).value = \
        f"=SUM(H{subtotal_row}:S{subtotal_row})"
    ws.cell(row=total_row, column=column_index_from_string("R")).value = \
        f"=H{total_row}-M{total_row}"
    adv = ident.get("预支现金")
    if adv is not None:
        ws.cell(row=total_row, column=column_index_from_string("M")).value = adv
    # S3 = 报销金额, 指向总额行 R 列
    ws["S3"] = f"=R{total_row}"

    # 自校验: 直接用台账金额按列汇总, 让上层无需依赖 LibreOffice 即可知道数值并核对
    col_sums = {}
    for it in items:
        col = resolve_col(it)
        amt = it.get("金额")
        if isinstance(amt, (int, float)):
            col_sums[col] = round(col_sums.get(col, 0) + amt, 2)
    grand = round(sum(col_sums.values()), 2)
    adv_val = ident.get("预支现金") or 0
    reimburse = round(grand - (adv_val if isinstance(adv_val, (int, float)) else 0), 2)

    wb.save(args.xlsx)
    print(json.dumps({
        "status": "ok", "xlsx": args.xlsx, "items": n,
        "data_rows": [DATA_START, last_data_row],
        "subtotal_row": subtotal_row, "total_row": total_row,
        "computed": {"列小计": col_sums, "费用合计H40": grand,
                      "预支现金M40": adv_val, "费用报销金额R40": reimburse},
        "note": "computed 为脚本按台账直接加总的预期值; 跑 recalc 后表内公式结果应与之逐项一致。"
                " recalc 首选 xlsx skill 的 recalc.py(需 LibreOffice); 无 LibreOffice 时改用同目录 recalc_fallback.py。"
    }, ensure_ascii=False))

if __name__ == "__main__":
    main()
